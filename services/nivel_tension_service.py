"""
Servicio para Cambio de Nivel de Tensión (NT1 → NT2).

Replica la lógica de la hoja "Resumen Final N2" del archivo NT2 Calculadora:
  K = ahorro tarifa por OR (lookup en tabla TARIFAS_OR)
  L = 100 * energía  (ahorro tarifario fijo NT1→NT2)
  M = K * energía
  N = L + M           → AHORRO MENSUAL (Distribución + Tarifa)
  Q = 1.1 * inversión (precio venta = margen 10%)
  R = Q / N           → ROI en meses
  Y = max(L, M, N)
  AA = (120 - 6) * Y  → AHORRO VIDA ÚTIL del proyecto (10 años - 6 meses financiación)

Genera el PDF (vía Slides), guarda PPTX renombrado en Shared Drive, mantiene
Excel acumulado con cada simulación, y devuelve los bytes del PDF para el
borrador de Gmail.
"""

import io
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.metabase_service import consultar_hubspot_general
from services.slides_service import generar_simulacion_desde_template
from services.drive_service import (
    subir_pdf,
    subir_bytes,
    listar_archivos_por_nombre,
    descargar_archivo,
    actualizar_archivo,
)


def _nt2_creds() -> str | None:
    """Credenciales de la SA del proyecto NT2SLIDES (con Slides API habilitada)."""
    return os.getenv("NT2_CREDENTIALS_PATH")

# ---------------------------------------------------------------------------
# TABLA TARIFAS — viene de la hoja "Resumen" col D ("Diferencia Tarifas OR-Bia")
# del archivo NT2 Calculadora. Actualizar acá si Bia cambia esta tabla.
# ---------------------------------------------------------------------------
TARIFAS_OR = {
    "AFINIA CARIBE_MAR":       39,
    "AIRE CARIBE_SOL":         87,
    "CEDENAR NARIÑO":          52,
    "CELSIA_TOLIMA TOLIMA":    13,
    "CEO CAUCA":              130,
    "EMCALI CALI":             22,
    "ENEL CUNDINAMARCA":      -24,
    "EPM ANTIOQUIA":          -27,
    "ESSA SANTANDER":          61,
}

# Constantes del modelo financiero (de las celdas hardcoded en row 3 del Excel)
AHORRO_NT1_NT2_KWH   = 100        # J3 — ahorro tarifario por pasar de NT1 a NT2
MARGEN_VENTA         = 1.1        # P3 — margen sobre costo (10%)
PERIODOS_FINANC      = 6          # V3 — periodos de financiación (meses)
VIDA_UTIL_MESES      = 120        # 10 años

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def operadores_red() -> list[str]:
    """Lista ordenada de ORs para el dropdown del UI."""
    return sorted(TARIFAS_OR.keys())


def calcular_simulacion(energia_kwh: float, inversion: float, operador_red: str) -> dict:
    """Aplica las fórmulas de la calculadora NT2 y devuelve los 4 valores del PDF."""
    if operador_red not in TARIFAS_OR:
        raise ValueError(f"Operador de Red '{operador_red}' no está en la tabla de tarifas")
    if energia_kwh <= 0:
        raise ValueError("La energía debe ser mayor a 0 kWh")
    if inversion <= 0:
        raise ValueError("La inversión debe ser mayor a 0")

    k = TARIFAS_OR[operador_red]
    l = AHORRO_NT1_NT2_KWH * energia_kwh
    m = k * energia_kwh
    n = l + m                                       # Ahorro mensual total
    q = MARGEN_VENTA * inversion                    # Precio venta
    roi_meses = q / n if n > 0 else float("inf")    # ROI a venta
    y = max(l, m, n)
    vida_util = (VIDA_UTIL_MESES - PERIODOS_FINANC) * y

    return {
        "ahorro_mensual":   n,
        "inversion":        inversion,
        "vida_util":        vida_util,
        "roi_meses":        roi_meses,
        # Auxiliares por trazabilidad
        "energia_kwh":      energia_kwh,
        "operador_red":     operador_red,
        "tarifa_or":        k,
        "ahorro_nt1_nt2":   l,
        "ahorro_tarifa":    m,
        "precio_venta":     q,
    }


def formato_cop(valor: float) -> str:
    """Formato '$1.234.567' usado en el PDF (miles con punto)."""
    return f"${valor:,.0f}".replace(",", ".")


def formato_meses(valor: float) -> str:
    """Formato '19,7 meses' (decimal con coma)."""
    return f"{valor:,.1f} meses".replace(".", ",")


def slug_archivo(razon_social: str, codigo_bia: str) -> str:
    """Nombre limpio para los archivos en Drive."""
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in razon_social).strip()
    return f"Simulacion NT2 - {safe} ({codigo_bia})"


# ---------------------------------------------------------------------------
# EXCEL ACUMULADO — se mantiene en la carpeta del Shared Drive y crece con
# una fila por cada simulación generada.
# ---------------------------------------------------------------------------
LOG_HEADERS = [
    "Fecha",
    "Código Bia",
    "Razón Social",
    "Dirección",
    "Ciudad",
    "Operador de Red",
    "Energía (kWh)",
    "Inversión (COP)",
    "Ahorro mensual (COP)",
    "Ahorro vida útil (COP)",
    "ROI (meses)",
]


def _aplicar_estilo_header(ws, ncols: int):
    fill   = PatternFill("solid", fgColor="1F4E79")
    font_h = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    align  = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill      = fill
        cell.font      = font_h
        cell.alignment = align
        cell.border    = border


def _autofit(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter  = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 4
    ws.row_dimensions[1].height = 20


def _crear_excel_log_nuevo() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulaciones"
    ws.append(LOG_HEADERS)
    _aplicar_estilo_header(ws, len(LOG_HEADERS))
    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _agregar_fila_al_log(excel_bytes: bytes, fila: list) -> bytes:
    """
    Reescribe el Excel desde cero: lee solo las filas que tienen datos reales,
    descarta filas vacías (que Google Sheets a veces inserta al auto-guardar
    cuando alguien abre el .xlsx en línea), agrega la nueva fila al final.

    Esto evita que el append "se hunda" en una sábana de filas vacías y queda
    invisible para el usuario.
    """
    wb_existente = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws_existente = (
        wb_existente["Simulaciones"]
        if "Simulaciones" in wb_existente.sheetnames
        else wb_existente.worksheets[0]
    )

    ncols = len(LOG_HEADERS)
    filas_con_datos = []
    for r in range(2, ws_existente.max_row + 1):
        valores = [ws_existente.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        if any(v is not None and (not isinstance(v, str) or v.strip() != "") for v in valores):
            filas_con_datos.append(valores)

    # Construir un workbook limpio desde cero
    wb_nuevo = openpyxl.Workbook()
    ws_nuevo = wb_nuevo.active
    ws_nuevo.title = "Simulaciones"
    ws_nuevo.append(LOG_HEADERS)
    _aplicar_estilo_header(ws_nuevo, ncols)
    for f in filas_con_datos:
        ws_nuevo.append(f)
    ws_nuevo.append(fila)
    _autofit(ws_nuevo)

    buf = io.BytesIO()
    wb_nuevo.save(buf)
    return buf.getvalue()


def actualizar_log_excel(parent_drive_id: str, fila: list) -> str:
    """Busca el Excel log en la carpeta; si no existe, lo crea. Devuelve file id.

    El lookup es por nombre solamente (sin filtrar por mimetype) para tolerar
    que Drive a veces almacena el XLSX con mimetype ligeramente distinto.
    """
    nombre_log = os.getenv("NIVEL_TENSION_LOG_FILENAME", "Simulaciones Nivel de Tensión.xlsx")
    creds      = _nt2_creds()
    existentes = listar_archivos_por_nombre(nombre_log, parent_drive_id, creds_path=creds)

    if existentes:
        file_id     = existentes[0]["id"]
        actual      = descargar_archivo(file_id, creds_path=creds)
        actualizado = _agregar_fila_al_log(actual, fila)
        actualizar_archivo(file_id, actualizado, MIME_XLSX, creds_path=creds)
        return file_id

    nuevo      = _crear_excel_log_nuevo()
    nuevo_con  = _agregar_fila_al_log(nuevo, fila)
    return subir_bytes(nombre_log, nuevo_con, parent_drive_id, MIME_XLSX, creds_path=creds)


# ---------------------------------------------------------------------------
# ORQUESTACIÓN COMPLETA — usada por la página Streamlit
# ---------------------------------------------------------------------------
async def procesar_simulacion(codigo_bia: str, inversion: float, operador_red: str) -> dict:
    """
    Flujo completo:
      1. Consulta Hubspot General → razón social, dirección, ciudad, energía.
      2. Calcula los 4 valores con la lógica NT2.
      3. Genera PDF + PPTX a partir de la plantilla en Slides.
      4. Sube PPTX renombrado al Shared Drive Nivel de Tensión.
      5. Agrega fila al Excel acumulado.

    Retorna dict con pdf_bytes, datos del cliente, valores calculados y URL del Drive.
    """
    parent_id   = os.getenv("NIVEL_TENSION_PARENT_ID")
    template_id = os.getenv("NIVEL_TENSION_TEMPLATE_ID")
    if not parent_id or not template_id:
        raise ValueError("Faltan NIVEL_TENSION_PARENT_ID o NIVEL_TENSION_TEMPLATE_ID en .env")

    cliente = await consultar_hubspot_general(codigo_bia)
    sim     = calcular_simulacion(cliente["energia"], inversion, operador_red)

    placeholders = {
        "{{RAZON_SOCIAL}}":   cliente["razon_social"] or "[Razón social pendiente]",
        "{{AHORRO_MENSUAL}}": formato_cop(sim["ahorro_mensual"]),
        "{{INVERSION}}":      formato_cop(sim["inversion"]),
        "{{VIDA_UTIL}}":      formato_cop(sim["vida_util"]),
        "{{ROI_MESES}}":      formato_meses(sim["roi_meses"]),
    }

    nombre = slug_archivo(cliente["razon_social"] or codigo_bia, codigo_bia)
    pdf_bytes, slides_url = generar_simulacion_desde_template(
        template_id   = template_id,
        parent_id     = parent_id,
        nombre_copia  = nombre,
        placeholders  = placeholders,
    )

    fila = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        codigo_bia,
        cliente["razon_social"],
        cliente["direccion"],
        cliente["ciudad"],
        operador_red,
        sim["energia_kwh"],
        sim["inversion"],
        round(sim["ahorro_mensual"], 0),
        round(sim["vida_util"], 0),
        round(sim["roi_meses"], 2),
    ]
    actualizar_log_excel(parent_id, fila)

    return {
        "pdf":         pdf_bytes,
        "pdf_nombre":  f"{nombre}.pdf",
        "slides_url":  slides_url,
        "cliente":     cliente,
        "simulacion":  sim,
        "drive_url":   f"https://drive.google.com/drive/folders/{parent_id}",
    }
