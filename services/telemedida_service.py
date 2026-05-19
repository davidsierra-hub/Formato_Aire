import io
import openpyxl


def obtener_ips_disponibles(lineas_file, cantidad: int) -> list:
    """
    Reads the uploaded Líneas Excel file.
    Returns up to `cantidad` unique IPs where APN=bia.claro.com.co
    and Cod. Interno BIA=Disponible.
    """
    contenido = lineas_file.read()
    lineas_file.seek(0)

    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)

    # Try sheet named "Líneas", fallback to first sheet
    if "Líneas" in wb.sheetnames:
        ws = wb["Líneas"]
    elif "Lineas" in wb.sheetnames:
        ws = wb["Lineas"]
    else:
        ws = wb.active

    # Columns: A=ICCID, B=Dirección IP, C=NUMERO, D=Plan, E=IMSI, F=APN, G=Telemedida, H=Cod. Interno BIA
    ips = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        ip = str(row[1] or "").strip()
        apn = str(row[5] or "").strip().lower()
        cod_interno = str(row[7] or "").strip().lower()

        if apn == "bia.claro.com.co" and cod_interno == "disponible" and ip and ip not in seen:
            ips.append(ip)
            seen.add(ip)
            if len(ips) >= cantidad:
                break

    return ips
