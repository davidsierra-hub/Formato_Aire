import io
import os
import zipfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

from services.wms_serial_service import consultar_wms_por_seriales
from services.telemedida_service import obtener_ips_disponibles
from services.wms_service import descargar_pdf
from services.drive_service import crear_o_buscar_carpeta, subir_bytes, subir_pdf

MESES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def nombre_carpeta_hoy() -> str:
    hoy = datetime.now()
    return f"Medidores directa {hoy.day}_{MESES[hoy.month]}_{hoy.year}"


def generar_excel(seriales: list, datos_wms: dict, ips: list) -> bytes:
    from openpyxl.styles import PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medidores Directa IP"

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    DATA_FONT   = Font(name="Calibri", size=11)
    CENTER      = Alignment(horizontal="center", vertical="center")
    THIN        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    headers = ["Nombre", "Serial", "Marca", "IP"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font        = HEADER_FONT
        cell.fill        = HEADER_FILL
        cell.alignment   = CENTER
        cell.border      = BORDER

    for i, serial in enumerate(seriales):
        dato    = datos_wms.get(serial, {})
        values  = [serial, serial, dato.get("marca") or "", ips[i] if i < len(ips) else ""]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i + 2, column=col, value=val)
            cell.font      = DATA_FONT
            cell.alignment = CENTER
            cell.border    = BORDER

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    ws.row_dimensions[1].height = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


async def obtener_pdfs_calibracion(seriales: list, datos_wms: dict) -> dict:
    """Returns {serial: pdf_bytes} for all serials with a calibration cert URL."""
    pdfs = {}
    for serial in seriales:
        url = datos_wms.get(serial, {}).get("certificado_calibracion")
        if url:
            pdfs[serial] = await descargar_pdf(url)
    return pdfs


def crear_zip(pdfs: dict) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for serial, pdf_bytes in pdfs.items():
            zf.writestr(f"{serial}.pdf", pdf_bytes)
    zip_buffer.seek(0)
    return zip_buffer.read()


async def obtener_certificado_conformidad(seriales: list, datos_wms: dict) -> bytes | None:
    for serial in seriales:
        url = datos_wms.get(serial, {}).get("certificado_conformidad")
        if url:
            return await descargar_pdf(url)
    return None


def subir_a_drive(excel_bytes: bytes, conformidad_bytes: bytes | None, pdfs_calibracion: dict) -> str:
    """
    Creates dated folder inside MEDIDORES_DIRECTA_PARENT_ID, uploads:
    - Medidores Directa IP.xlsx
    - Certificado Conformidad.pdf
    - Subfolder 'Certificados calibración' with one PDF per serial
    Returns the Drive folder URL.
    """
    parent_id = os.getenv("MEDIDORES_DIRECTA_PARENT_ID")
    if not parent_id:
        raise ValueError("MEDIDORES_DIRECTA_PARENT_ID no configurado en .env")

    nombre = nombre_carpeta_hoy()
    folder_id = crear_o_buscar_carpeta(nombre, parent_id)

    subir_bytes("Medidores Directa IP.xlsx", excel_bytes, folder_id, MIME_XLSX)

    if conformidad_bytes:
        subir_pdf("Certificado Conformidad.pdf", conformidad_bytes, folder_id)

    if pdfs_calibracion:
        cal_folder_id = crear_o_buscar_carpeta("Certificados calibración", folder_id)
        for serial, pdf_bytes in pdfs_calibracion.items():
            subir_pdf(f"{serial}.pdf", pdf_bytes, cal_folder_id)

    return f"https://drive.google.com/drive/folders/{folder_id}"


async def procesar_medidores_directa(seriales: list, lineas_file) -> dict:
    datos_wms = await consultar_wms_por_seriales(seriales)
    ips = obtener_ips_disponibles(lineas_file, len(seriales))

    excel_bytes = generar_excel(seriales, datos_wms, ips)
    pdfs_calibracion = await obtener_pdfs_calibracion(seriales, datos_wms)
    conformidad_bytes = await obtener_certificado_conformidad(seriales, datos_wms)
    zip_bytes = crear_zip(pdfs_calibracion)

    return {
        "excel": excel_bytes,
        "zip": zip_bytes,
        "conformidad": conformidad_bytes,
        "pdfs_calibracion": pdfs_calibracion,
        "seriales_encontrados": len(datos_wms),
        "seriales_no_encontrados": [s for s in seriales if s not in datos_wms],
        "ips_asignadas": len(ips),
    }
