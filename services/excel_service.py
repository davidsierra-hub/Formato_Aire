from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

NO_FILL = PatternFill(fill_type=None)
YELLOW  = PatternFill(fill_type="solid", fgColor="FFFF00")


def clear_yellow(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
                if "FFFF00" in cell.fill.fgColor.rgb:
                    cell.fill = NO_FILL


def fill_row(sheet, row_num, marca, serial):
    sheet[f"C{row_num}"] = marca
    sheet[f"D{row_num}"] = str(serial)


def procesar_excel(template_bytes: bytes, datos: dict) -> bytes:
    nic          = datos["nic"]
    razon_social = datos["razon_social"]
    tipo_medida  = datos["tipo_medida"]
    direccion    = datos["direccion"]
    departamento = datos["departamento"]
    ciudad       = datos["ciudad"]
    equipos      = datos["equipos"]

    medidores = [e for e in equipos if "medidor" in (e["categoria"] or "").lower()]
    tcs       = [e for e in equipos if "transformador de corriente" in (e["categoria"] or "").lower()]
    tps       = [e for e in equipos if "transformador de potencial" in (e["categoria"] or "").lower()]

    wb = load_workbook(io.BytesIO(template_bytes))

    # Hoja principal — C24 y C25 no se llenan intencionalmente
    s = wb["Solicitud Acompañamiento al OR"]
    s["C28"] = razon_social
    s["C29"] = direccion
    s["C30"] = ciudad
    s["C31"] = departamento
    clear_yellow(s)

    # Sistema de medida
    sm = wb["Sistema de medida"]
    if medidores:
        fill_row(sm, 18, medidores[0]["marca"], medidores[0]["serial"])
    for i, tc in enumerate(tcs[:3]):
        fill_row(sm, 20 + i, tc["marca"], tc["serial"])
    for i, tp in enumerate(tps[:3]):
        fill_row(sm, 23 + i, tp["marca"], tp["serial"])
    sm["C26"] = tipo_medida
    clear_yellow(sm)

    # Formato comunicaciones
    fc = wb["Formato comunicaciones"]
    clear_yellow(fc)
    medidor = medidores[0] if medidores else {}
    fc["K15"]  = nic
    fc["AN15"] = direccion
    fc["K18"]  = razon_social
    fc["G26"]  = str(medidor.get("serial", ""))
    fc["P26"]  = medidor.get("marca", "")
    fc["AA26"] = "0,5s"
    fc["AI26"] = tcs[0]["sku"] if tcs else ""
    fc["AR26"] = tps[0]["sku"] if tps else ""
    for cell in ("G26", "P26", "AI26", "AR26"):
        fc[cell].fill = NO_FILL
    fc["V26"].fill  = YELLOW  # conservar amarillo
    fc["Q41"].fill  = YELLOW  # conservar amarillo

    # Hojas EDT: conservar solo la que corresponde al tipo de medida
    # Evaluar semidirecta antes que indirecta (substring "directa" está en ambas)
    tipo_lower = tipo_medida.lower()
    edt_map = [
        ("semidirecta", "EDT Semi",      ["EDT Directa", "EDT Indirecta"]),
        ("indirecta",   "EDT Indirecta", ["EDT Directa",  "EDT Semi"]),
        ("directa",     "EDT Directa",   ["EDT Semi",     "EDT Indirecta"]),
    ]
    for key, keep, remove in edt_map:
        if key in tipo_lower:
            for sname in remove:
                if sname in wb.sheetnames:
                    del wb[sname]
            if keep in wb.sheetnames:
                wb[keep].title = "EDT"
            break

    # Limpiar amarillo en hojas restantes
    hojas_principales = {
        "Solicitud Acompañamiento al OR",
        "Sistema de medida",
        "Formato comunicaciones",
    }
    for shname in wb.sheetnames:
        if shname not in hojas_principales:
            clear_yellow(wb[shname])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
